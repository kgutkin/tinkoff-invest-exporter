import sqlite3

from openpyxl import load_workbook

MIN_ROW = 9
MAX_ROW = 1000
MAX_COL = 85
DB_NAME = 'Transactions.db'
SCHEMA_FILE_NAME = './files/schema.sql'
TRANSACTION_TYPE_COLUMN = 6


def export_header_columns_indexes(sheet, connection):
    headers = None
    for row in sheet.iter_rows(min_row=MIN_ROW - 1, max_row=MIN_ROW - 1, min_col=1, max_col=MAX_COL):
        headers = [(i, str(col.value).replace('\n', '')) for i, col in enumerate(row) if col.value is not None]
        break

    cursor = connection.cursor()

    for header in headers:
        cursor.execute('insert into transaction_headers (columnIndex, name) values(?,?)', header)

    connection.commit()

    return headers


def init_db_schema(connection):
    sql = open(SCHEMA_FILE_NAME, 'r').read()
    cursor = connection.cursor()
    for cmd in sql.split('\n\n'):
        cursor.execute(cmd)
    connection.commit()


def init_db_connection():
    connection = sqlite3.connect(DB_NAME)
    init_db_schema(connection)
    return connection


def insert_transaction(con, transactions):
    cursor = con.cursor()
    sql = '''INSERT INTO transactions (
        id, errandNumber, conclusionDate, time, tradingPlatform, tradeRegime, type, assetShortName, ticker,
        price, priceCurrency, count, sumWithoutNkd, nkd, sum, transactionCurrency, commission, commissionCurrency, 
        repoRate, counterparty, settlementDay, deliveryDate, brokerStatus, contractType, contractNumber, contractDate)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)'''
    for i in [9, 12, 13, 14, 16, 18]:
        if (transactions[i] is not None):
            transactions[i] = float(transactions[i].replace(',', '.'))
    cursor.execute(sql, tuple(transactions))
    con.commit()


def export2db(file_name):
    workbook = load_workbook(filename=file_name, read_only=True)
    sheet = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])

    connection = init_db_connection()

    counter = 0
    error_counter = 0
    repo_counter = 0
    header_columns_indexes = [columnIndex for (columnIndex, name) in export_header_columns_indexes(sheet, connection)]

    for row in sheet.iter_rows(min_row=MIN_ROW, max_row=MAX_ROW, min_col=1, max_col=MAX_COL):
        try:
            if row is not None and row[0].value is not None:
                if "1.2" in str(row[0].value):
                    break
                transactions = [row[i].value for i in header_columns_indexes]

                if transactions[TRANSACTION_TYPE_COLUMN] in ['Покупка', 'Продажа']:
                    insert_transaction(connection, transactions)
                    counter += 1
                else:
                    repo_counter += 1
        except Exception:
            error_counter += 1

    connection.close()

    print(f'{"=" * 30}\nЭкспорт из экселя в sqlite3\n{"=" * 30}\nКоличество успешно обработаных транзакции: {counter}\n'
          f'Количество неизвестных сделок (например, РЕПО 1 Покупка/РЕПО 1 Продажа): {repo_counter}\n'
          f'Количество ошибок: {error_counter}\n\n')


def read_transactions_from_db():
    connection = sqlite3.connect(DB_NAME)
    cursor = connection.cursor()
    transactions = cursor.execute("select * from transactions").fetchall()
    connection.close()
    return transactions


def read_header_names_from_db():
    connection = sqlite3.connect(DB_NAME)
    cursor = connection.cursor()
    headers = cursor.execute("select * from transaction_headers").fetchall()
    connection.close()
    return [name for (_, name) in headers]


def store_settings_to_db(args):
    connection = init_db_connection()
    cursor = connection.cursor()
    sql = 'insert or replace into settings values '

    values = [f"('{key}', '{value}')" for (key, value) in args.items() if value is not None and key not in ['storeArgs', 'fileName']]
    if values:
        full_sql = sql + ', '.join(values) + ';'
        cursor.execute(full_sql)
        connection.commit()

    connection.close()


def read_settings_from_db():
    connection = sqlite3.connect(DB_NAME)
    cursor = connection.cursor()
    settings = cursor.execute("select * from settings").fetchall()
    connection.close()
    return {setting[0]: setting[1] for setting in settings}


if __name__ == '__main__':
    export2db("./files/broker_rep.xlsx")
